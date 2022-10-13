from openpyxl import load_workbook, Workbook

c_name = 1
c_id = 2
c_money = 3

default_money = 30000000

wb = load_workbook("./db/userDB.xlsx")
ws = wb.active

def loadFile():
    wb = load_workbook("./db/userDB.xlsx")
    ws = wb.active

def saveFile():
    wb.save("./db/userDB.xlsx")
    wb.close()

def checkUser(_name, _id):
    print("user.py - checkUser")
    print(str(_name) + "<" + str(_id) + ">의 존재 여부 확인")
    print("")

    loadFile()

    userNum = checkUserNum()
    print("등록된 유저수: ", userNum)
    print("")

    print("이름과 고유번호 탐색")
    print("")

    for row in range(2, 3+userNum):
        print(row, "번째 줄 name: ", ws.cell(row,c_name).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", ws.cell(row, c_name).value == _name)

        print(row,"번째 줄 id: ", ws.cell(row,c_id).value)
        print("입력된 id: ", hex(_id))
        print("고유번호정보와 일치 여부: ", ws.cell(row, c_id).value == hex(_id))
        print("")

        if ws.cell(row, c_name).value == _name and ws.cell(row,c_id).value == hex(_id):
            print("등록된  이름과 고유번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")
            print("")

            saveFile()

            return True, row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()
    print("발견 실패")

    return False, None

def checkUserNum():
    loadFile()

    count = 0

    for row in range(2, ws.max_row+1):
        if ws.cell(row,c_name).value != None:
            count = count+1
        else:
            count = count
    return count
    

def findRow(_name, _id):
    loadFile()

    print("등록된 유저수: ", checkUserNum())
    print("")

    print("이름과 고유번호 탐색")
    for row in range(2, ws.max_row+2):
        print(row, "번째 줄 name: ", ws.cell(row,c_name).value)
        print("입력된 name: ", _name)
        print("이름과 일치 여부: ", ws.cell(row, c_name).value == _name)

        print(row,"번째 줄 id: ", ws.cell(row,c_id).value)
        print("입력된 id: ", hex(_id))
        print("고유번호정보와 일치 여부: ", ws.cell(row, c_id).value == hex(_id))
        print("")

        if ws.cell(row, c_name).value == _name and ws.cell(row,c_id).value == hex(_id):
            print("등록된  이름과 고유번호를 발견")
            print("등록된  값의 위치: ",  row, "번째 줄")

            saveFile()

            return row
            break
        else:
            print("등록된 정보를 탐색 실패, 재탐색 실시")

    saveFile()

    return None

def getMoney(_name,_row):
    print("user.py - getMoney")
    loadFile()

    print(_name, "의 돈을 탐색")

    result = ws.cell(_row, c_money).value
    print(_name,"의 보유 자산: ", result)

    saveFile()

    return result

def remit(sender, sender_row, receiver, receiver_row, _amount):
    print("user.py - remit")
    #loadFile()
    
    print("보내는 사람: ", sender)
    print("받는 사람: ", receiver)
    print("보내는 돈: ", _amount)
    print("")

    modifyMoney(receiver, receiver_row, _amount)
    modifyMoney(sender, sender_row, -int(_amount))

    print("")

def modifyMoney(_target, _row, _amount):
    print("user.py - modifyMoney")
    loadFile()

    print(_target, "의 자산데이터 수정")
    print(_target, "의 자산: " + str(ws.cell(_row, c_money).value))
    print("추가할 액수: ", _amount)
    ws.cell(_row, c_money).value += int(_amount)

    print("자산데이터 수정 완료")
    print("수정된", _target, "의 자산: ", ws.cell(_row, c_money).value)
    
    saveFile()

def checkRow():
    loadFile()
    print("첫번쨰 빈 곳을 탐색")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row,1).value is None:
            return row
            break
    _result = ws.max_row+1
    saveFile()

    return _result

def signup(_name, _id):
    loadFile()
    print("첫번째 빈곳: ", checkRow())
    _row = checkRow()

    print("데이터 추가 시작")
    ws.cell(row=_row, column=c_name, value=_name)
    print("이름 추가 | ",  ws.cell(_row,c_name).value)
    ws.cell(row=_row, column=c_id, value =hex(_id))
    print("고유번호 추가 | ", ws.cell(_row,c_id).value)
    ws.cell(row=_row, column=c_money, value = default_money)
    print("기본 자금 지급 | ", ws.cell(_row,c_money).value)

    saveFile()

    print("데이터 추가 완료")

def userInfo(_name, _id):
    loadFile()

    userRow = findRow(_name, _id)

    if not userRow == None:
        print("사용자 정보 발견, 보유 자산과 레벨을 반환")
        print("보유자산: ", ws.cell(userRow,c_money).value)
        return ws.cell(userRow,c_money).value
    else:
        #사용자 정보 없음
        return None


def delete():
    loadFile()

    print("유저 데이터를 삭제")

    ws.delete_rows(2,ws.max_row)
    saveFile()

    print("데이터 삭제 완료")